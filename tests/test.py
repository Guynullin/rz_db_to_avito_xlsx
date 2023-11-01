import sys
import os
sys.path.append(os.getcwd())

from test_data import alloy_card, forged_card
from xlsx_api.format_title import format_alloy_title, format_forged_title
from openpyxl import load_workbook
import os
import json
from config import XLSX_PATH, FILES_LIST

def test_format_titles():
    assert format_alloy_title(alloy_card) == "Литой диск KOKO KUTURE R17 4x100"
    assert format_forged_title(forged_card) == "Кованый диск RZ R22 6X139.7"

def test_equals_file_rows_id():
    rows_dict = {}
    for root, dirs, files in os.walk(XLSX_PATH):  
        for filename in files:
            file_path = os.path.join(XLSX_PATH, filename)
            assert os.path.exists(file_path) 
            assert os.path.isfile(file_path)
            assert filename in FILES_LIST
            rows_dict[filename] = []
            wb = load_workbook(file_path)
            data_list = []
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                id = ws[3][1].value
                max_id = ws[ws.max_row][1].value
                data = {sheetname : {'id' : id, 'max_id' : max_id}}
                data_list.append(data)
            rows_dict[filename] = json.dumps(data_list)    
            wb.close()
    
    assert len(rows_dict) > 0
    for filename in rows_dict:
        data = rows_dict[filename]
        for key in rows_dict:
            assert rows_dict[key] == data


import os
import logging
import shutil
import copy
from openpyxl import load_workbook

from xlsx_api.record_forged_rims import record_forged_rims
from xlsx_api.record_alloy_rims import record_alloy_rims
from xlsx_api.record_tires import record_tires
from xlsx_api.record_springs import record_springs
from info_bot import send_message
from config import CITY_DATA, WEB_PATH, FILES_LIST


def parse_xlsx(input_file_path: str, db_cards: dict):

    for root, dirs, files in os.walk(input_file_path):  
        for filename in files:
            if filename in FILES_LIST:
                file_path = os.path.join(input_file_path, filename)

                if os.path.exists(file_path) and os.path.isfile(file_path):
                    wb = load_workbook(file_path)
                else:
                    send_message(f'ERROR\nfile path: {file_path} not exists')
                    logging.error(f'file path: {file_path} not exists')
                    return 0
                cards = copy.deepcopy(db_cards)
                for sheetname in wb.sheetnames:

                    ws = wb[sheetname]
                    if sheetname == 'диски кованые':
                        if record_forged_rims(cards=cards['forged'], ws=ws, city_data=CITY_DATA[filename]) != 1:
                            send_message('ERROR\nЛист кованные диски не записан!')
                    elif sheetname == 'диски литые':
                        if record_alloy_rims(cards=cards['alloy'], ws=ws, city_data=CITY_DATA[filename]) != 1:
                            send_message('ERROR\nЛист литые диски не записан!')
                    elif sheetname == 'шины':
                        if record_tires(cards=cards['tires'], ws=ws, city_data=CITY_DATA[filename]) != 1:
                            send_message('ERROR\nЛист шины не записан!')
                    elif sheetname == 'пружины':
                        if record_springs(cards=cards['springs'], ws=ws, city_data=CITY_DATA[filename]) != 1:
                            send_message('ERROR\nЛист пружины не записан!')

                wb.save(file_path)
                web_path = os.path.join(WEB_PATH, filename)
                shutil.copyfile(file_path, web_path)

                logging.info(f'The file was saved successfully. Path: {file_path}')
    
    return 1

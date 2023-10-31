import logging
import random
from openpyxl import worksheet

from xlsx_api.format_title import format_forged_title
from xlsx_api.format_desc import format_rim_desc



def record_forged_rims(cards: list, ws: worksheet, city_data: dict):
    cards_id_list = []
    title = 'Диски кованые'

    if ws[1][0].value == 'AvitoId' and ws[1][1].value == 'Id' and ws[1][2].value == 'Title'\
    and ws[1][3].value == 'Description' and ws[1][4].value == 'Price' and ws[1][5].value == 'RimDiameter'\
    and ws[1][6].value == 'RimBolts' and ws[1][7].value == 'RimBoltsDiameter' and ws[1][8].value == 'RimWidth'\
    and ws[1][9].value == 'RimOffset' and ws[1][10].value == 'RimDIA' and ws[1][11].value == 'RimType'\
    and ws[1][12].value == 'ProductType' and ws[1][13].value == 'ImageUrls' and ws[1][14].value == 'GoodsType'\
    and ws[1][15].value == 'AdType' and ws[1][19].value == 'TypeId' and ws[1][20].value == 'ManagerName'\
    and ws[1][21].value == 'Condition' and ws[1][22].value == 'AvitoStatus' and ws[1][23].value == 'ContactMethod'\
    and ws[1][24].value == 'Category' and ws[1][25].value == 'ListingFee' and ws[1][26].value == 'CompanyName'\
    and ws[1][27].value == 'DateBegin' and ws[1][16].value == 'Address' and ws[1][17].value == 'EMail'\
    and ws[1][18].value == 'ContactPhone':  
        
        logging.info(f'<Record forged rims> The number of rows from db: {len(cards)}')

        # collect cards id
        for card in cards:
            cards_id_list.append(card['id'])

        del_rows = []
        for index, row in enumerate(ws.rows):
            if row[1].value == 'Id':
                continue
            if row[1].value != None:
                if row[1].value not in cards_id_list:
                    del_rows.append(index+1)
                else:
                    for card in cards:
                        if index == 3:
                                print(card)
                        if card['id'] == row[1].value:
                            if row[2].value == None:
                                title = format_forged_title(card)
                                row[2].value = title
                            if row[3].value == None:
                                title = format_forged_title(card)
                                row[3].value = format_rim_desc(title=title)
                            if row[4].value != card['price']:
                                row[4].value = card['price']
                            if row[5].value != card['diameter']:
                                row[5].value = card['diameter']
                            if row[6].value != card['bolts']:
                                row[6].value = card['bolts']
                            if row[7].value != card['pcd']:
                                row[7].value = card['pcd']
                            if row[8].value != card['width']:
                                row[8].value = card['width']
                            if row[9].value != card['et']:
                                row[9].value = card['et']
                            if row[10].value != card['dia']:
                                row[10].value = card['dia']
                            if row[11].value != 'Кованые':
                                row[11].value = 'Кованые'
                            if row[12].value != 'Диски':
                                row[12].value = 'Диски'
                            if row[13].value == None:
                                photo_list = []
                                if 'live_photo' in card:
                                    photo_list = ['https://rimzona.ru/media/{}'.format(slug) for slug in card['live_photo']] 
                                if 'photo' in card:
                                    photo_list += ['https://rimzona.ru/media/{}'.format(slug) for slug in card['photo']]
                                row[13].value = ' | '.join(photo_list)
                            if row[14].value != 'Шины, диски и колёса':
                                row[14].value = 'Шины, диски и колёса'
                            if row[15].value != 'Товар от производителя':
                                row[15].value = 'Товар от производителя'
                            if row[16].value == None:
                                row[16].value = random.choice(city_data['adress'])
                            if row[17].value == None:
                                row[17].value = city_data['email']
                            if row[18].value == None:
                                row[18].value = city_data['phone']
                            if row[19].value == None:
                                row[19].value = '10-046'
                            if row[20].value == None:
                                row[20].value = 'Rimzona'
                            if row[21].value == None:
                                row[21].value = 'Новое'
                            if row[22].value == None:
                                row[22].value = 'Активно'
                            if row[23].value == None:
                                row[23].value = 'По телефону и в сообщениях'
                            if row[24].value == None:
                                row[24].value = 'Запчасти и аксессуары'
                            if row[25].value == None:
                                row[25].value = 'Package'
                            if row[26].value == None:
                                row[26].value = 'Rimzona : сеть магазинов стильных дисков и шин'
                            cards.remove(card)
                            break
            else:
                del_rows.append(index+1)

        # remove old and empty rows 
        logging.info(f'<Record forged rims> The number of rows to be deleted: {len(del_rows)}')
        for i in reversed(del_rows):
            ws.delete_rows(i)
        
        if len(cards) > 0:
            logging.info(f"<Record forged rims> Last line number: {ws.max_row}")
            logging.info(f'<Record forged rims> The number of lines to be added {len(cards)}')
            for card in cards:
                index = ws.max_row + 1
                if ws[index][1].value == None:
                    ws[index][1].value = card['id']
                    title = format_forged_title(card)
                    ws[index][2].value = title
                    ws[index][3].value = format_rim_desc(title=title)
                    ws[index][4].value = card['price']
                    ws[index][5].value = card['diameter']
                    ws[index][6].value = card['bolts']
                    ws[index][7].value = card['pcd']
                    ws[index][8].value = card['width']
                    ws[index][9].value = card['et']
                    ws[index][10].value = card['dia']
                    ws[index][11].value = 'Кованые'
                    ws[index][12].value = 'Диски'
                    photo_list = []
                    if 'live_photo' in card:
                        photo_list = ['https://rimzona.ru/media/{}'.format(slug) for slug in card['live_photo']] 
                    if 'photo' in card:
                        photo_list += ['https://rimzona.ru/media/{}'.format(slug) for slug in card['photo']]
                    ws[index][13].value = ' | '.join(photo_list)
                    ws[index][14].value = 'Шины, диски и колёса'
                    ws[index][15].value = 'Товар от производителя'
                    ws[index][16].value = random.choice(city_data['adress'])
                    ws[index][17].value = city_data['email']
                    ws[index][18].value = city_data['phone']
                    ws[index][19].value = '10-046'
                    ws[index][20].value = 'Rimzona'
                    ws[index][21].value = 'Новое'
                    ws[index][22].value = 'Активно'
                    ws[index][23].value = 'По телефону и в сообщениях'
                    ws[index][24].value = 'Запчасти и аксессуары'
                    ws[index][25].value = 'Package'
                    ws[index][26].value = 'Rimzona : сеть магазинов стильных дисков и шин'
                    logging.debug(f"{index} : {ws[index][1].value} : {ws[index][9].value}")

        logging.info(f"<Record forged rims> -END- Last line number: {ws.max_row}")
        return 1
    
    else:
        raise Exception('Лист кованые диски содержит некорректный титул')



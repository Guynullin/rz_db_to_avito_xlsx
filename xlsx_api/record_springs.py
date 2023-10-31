import logging
import random
from openpyxl import worksheet

from xlsx_api.format_desc import format_spring_desc


def record_springs(cards: list, ws: worksheet, city_data: dict):
    cards_id_list = []
    title = 'Пружины'
    

    if ws[1][0].value == 'AvitoId' and ws[1][1].value == 'Id' and ws[1][2].value == 'ContactPhone'\
        and ws[1][3].value == 'ListingFee' and ws[1][4].value == 'AvitoDateEnd' and ws[1][5].value == 'ProductType'\
        and ws[1][6].value == 'Price' and ws[1][7].value == 'Description' and ws[1][8].value == 'CompanyName'\
        and ws[1][9].value == 'Title' and ws[1][10].value == 'ContactMethod' and ws[1][11].value == 'Category'\
        and ws[1][12].value == 'GoodsType' and ws[1][13].value == 'AdType' and ws[1][14].value == 'ImageUrls'\
        and ws[1][15].value == 'TypeId' and ws[1][16].value == 'Condition' and ws[1][17].value == 'EMail'\
        and ws[1][18].value == 'Address' and ws[1][19].value == 'AvitoStatus' and ws[1][20].value == 'Brand'\
        and ws[1][21].value == 'Model' and ws[1][22].value == 'Availability' and ws[1][23].value == 'Generation'\
        and ws[1][24].value == 'SparePartType' and ws[1][25].value == 'Make' and ws[1][26].value == 'Originality'\
        and ws[1][27].value == 'OEM':

        logging.info(f'<Record springs> The number of rows from db: {len(cards)}')

        # collect cards id
        for card in cards:
            cards_id_list.append(card['id'])

        # print(f"tire len : {len(cards)}")
        # return

        del_rows = []
        for index, row in enumerate(ws.rows):
            if row[1].value == 'Id':
                continue
            if row[1].value != None:
                if row[1].value not in cards_id_list:
                    del_rows.append(index+1)
                else:
                    for card in cards:
                        if index == 3:
                                print(card)
                        if card['id'] == row[1].value:
                            if row[2].value == None:
                                row[2].value = city_data['phone']
                            if row[3].value == None:
                                row[3].value = 'Package'
                            # if row[4].value == None:
                            #     row[4].value = 'AvitoDateEnd'
                            if row[5].value == None:
                                row[5].value = 'Для автомобилей'
                            if row[6].value != card['price']:
                                row[6].value = card['price']
                            if row[7].value == None:
                                row[7].value = format_spring_desc(title=card['title'])
                            if row[8].value == None:
                                row[8].value = 'Rimzona : сеть магазинов стильных дисков и шин'
                            if row[9].value == None:
                                row[9].value = card['title']
                            if row[10].value == None:
                                row[10].value = 'По телефону и в сообщениях'
                            if row[11].value == None:
                                row[11].value = 'Запчасти и аксессуары'
                            if row[12].value == None:
                                row[12].value = 'Запчасти'
                            if row[13].value == None:
                                row[13].value = 'Товар от производителя'
                            if row[14].value == None:
                                photo_list = []
                                if 'live_photo' in card:
                                    photo_list = ['https://rimzona.ru/media/{}'.format(slug) for slug in card['live_photo']] 
                                if 'photo' in card:
                                    photo_list += ['https://rimzona.ru/media/{}'.format(slug) for slug in card['photo']]
                                row[14].value = ' | '.join(photo_list)
                            if row[15].value == None:
                                row[15].value = '11-623'
                            if row[16].value == None:
                                row[16].value = 'Новое'
                            if row[17].value == None:
                                row[17].value = city_data['email']
                            if row[18].value == None:
                                row[18].value = random.choice(city_data['adress'])
                            if row[19].value == None:
                                row[19].value = 'Активно'
                            if row[20].value != card['brand']:
                                row[20].value = card['brand']
                            if row[21].value == None:
                                row[21].value = card['model']
                            if row[22].value == None:
                                row[22].value = 'В наличии'
                            # WARN (Generation)
                            # if row[23].value == None:
                                # row[23].value = None
                            if row[24].value == None:
                                row[24].value = 'Подвеска'
                            if row[25].value == None:
                                row[25].value = card['make']
                            if row[26].value == None:
                                row[26].value = 'Оригинал'
                            
                            cards.remove(card)
                            break
            else:
                del_rows.append(index+1)

        # remove old and empty rows
        logging.info(f'<Record springs> The number of rows to be deleted: {len(del_rows)}')
        for i in reversed(del_rows):
            ws.delete_rows(i)

        if len(cards) > 0:
            logging.info(f"<Record springs> Last line number: {ws.max_row}")
            logging.info(f'<Record springs> The number of lines to be added {len(cards)}')
            for card in cards:
                index = ws.max_row + 1
                if ws[index][1].value == None:
                    ws[index][1].value = card['id']
                    ws[index][2].value = city_data['phone']
                    ws[index][3].value = 'Package'
                    # ws[index][4].value = 'AvitoDateEnd'
                    ws[index][5].value = 'Для автомобилей'
                    ws[index][6].value = card['price']
                    ws[index][7].value = format_spring_desc(title=card['title'])
                    ws[index][8].value = 'Rimzona : сеть магазинов стильных дисков и шин'
                    ws[index][9].value = card['title']
                    ws[index][10].value = 'По телефону и в сообщениях'
                    ws[index][11].value = 'Запчасти и аксессуары'
                    ws[index][12].value = 'Запчасти'
                    ws[index][13].value = 'Товар от производителя'
                    photo_list = []
                    if 'live_photo' in card:
                        photo_list = ['https://rimzona.ru/media/{}'.format(slug) for slug in card['live_photo']] 
                    if 'photo' in card:
                        photo_list += ['https://rimzona.ru/media/{}'.format(slug) for slug in card['photo']]
                    ws[index][14].value = ' | '.join(photo_list)
                    ws[index][15].value = '11-623'
                    ws[index][16].value = 'Новое'
                    ws[index][17].value = city_data['email']
                    ws[index][18].value = random.choice(city_data['adress'])
                    ws[index][19].value = 'Активно'
                    ws[index][20].value = card['brand']
                    ws[index][21].value = card['model']
                    ws[index][22].value = 'В наличии'
                    # WARN (Generation)
                    # ws[index][23].value = None
                    ws[index][24].value = 'Подвеска'
                    ws[index][25].value = card['make']
                    ws[index][26].value = 'Оригинал'
                    logging.debug(f"{index} : {ws[index][1].value} : {ws[index][9].value}")

        logging.info(f"<Record tires> -END- Last line number: {ws.max_row}")
        return 1
    else:
        raise Exception('Лист пружины содержит некорректный титул')


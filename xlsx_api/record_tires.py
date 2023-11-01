import logging
import random
from openpyxl import worksheet

from xlsx_api.format_desc import format_tire_desc


def record_tires(cards: list, ws: worksheet, city_data: dict):
    
    if ws[1][0].value == 'AvitoId' and ws[1][1].value == 'Id' and ws[1][2].value == 'Title'\
        and ws[1][3].value == 'Description' and ws[1][4].value == 'Price' and ws[1][5].value == 'RimDiameter'\
        and ws[1][6].value == 'ProductType' and ws[1][7].value == 'ImageUrls' and ws[1][8].value == 'GoodsType'\
        and ws[1][9].value == 'AdType' and ws[1][10].value == 'Address' and ws[1][11].value == 'EMail'\
        and ws[1][12].value == 'ContactPhone' and ws[1][13].value == 'TypeId' and ws[1][14].value == 'Condition'\
        and ws[1][15].value == 'AvitoStatus' and ws[1][16].value == 'ContactMethod' and ws[1][17].value == 'Category'\
        and ws[1][18].value == 'ListingFee' and ws[1][19].value == 'CompanyName' and ws[1][20].value == 'Quantity'\
        and ws[1][21].value == 'TireType' and ws[1][22].value == 'TireAspectRatio' and ws[1][23].value == 'LoadIndex'\
        and ws[1][24].value == 'DifferentWidthTires' and ws[1][25].value == 'SpeedIndex' and ws[1][26].value == 'RunFlat'\
        and ws[1][27].value == 'Model' and ws[1][28].value == 'Brand' and ws[1][29].value == 'TireSectionWidth':

        logging.info(f'<Record tires> The number of rows from db: {len(cards)}')

        # collect cards id
        cards_id_list = []
        for card in cards:
            cards_id_list.append(card['id'])

        del_rows = []
        for index, row in enumerate(ws.rows):
            if row[1].value == 'Id':
                continue
            if row[1].value != None:
                if row[1].value not in cards_id_list:
                    del_rows.append(index+1)
                else:
                    for card in cards:
                        if card['id'] == row[1].value:
                            if row[2].value == None:
                                row[2].value = card['title']
                            if row[3].value == None:
                                row[3].value = format_tire_desc(season=card['tire_type'], title=card['title'])
                            if row[4].value != card['price']:
                                row[4].value = card['price']
                            if row[5].value != card['diameter']:
                                row[5].value = card['diameter']
                            if row[6].value != 'Легковые шины':
                                row[6].value = 'Легковые шины'
                            if row[7].value == None:
                                photo_list = []
                                if 'live_photo' in card:
                                    photo_list = ['https://rimzona.ru/media/{}'.format(slug) for slug in card['live_photo']] 
                                if 'photo' in card:
                                    photo_list += ['https://rimzona.ru/media/{}'.format(slug) for slug in card['photo']]
                                row[7].value = ' | '.join(photo_list)
                            if row[8].value != 'Шины, диски и колёса':
                                row[8].value = 'Шины, диски и колёса'
                            if row[9].value != 'Товар от производителя':
                                row[9].value = 'Товар от производителя'
                            if row[10].value == None:
                                row[10].value = random.choice(city_data['adress'])
                            if row[11].value == None:
                                row[11].value = city_data['email']
                            if row[12].value == None:
                                row[12].value = city_data['phone']
                            if row[13].value != '10-048':
                                row[13].value = '10-048'
                            if row[14].value == None:
                                row[14].value = 'Новое'
                            if row[15].value == None:
                                row[15].value = 'Активно'
                            if row[16].value == None:
                                row[16].value = 'По телефону и в сообщениях'
                            if row[17].value == None:
                                row[17].value = 'Запчасти и аксессуары'
                            if row[18].value == None:
                                row[18].value = 'Package'
                            if row[19].value == None:
                                row[19].value = 'Rimzona : сеть магазинов стильных дисков и шин'
                            if row[20].value == None:
                                row[20].value = 'за 1 шт.'
                            if row[21].value == None:
                                row[21].value = card['tire_type']
                            if row[22].value != card['height']:
                                row[22].value = card['height']
                            if row[23].value != card['load_index']:
                                row[23].value = card['load_index']
                            if row[24].value == None:
                                row[24].value = 'Нет'
                            if row[25].value != card['speed_index']:
                                row[25].value = card['speed_index']
                            if row[26].value == None or row[26].value not in ['Да', 'Нет']:
                                row[26].value = card['runflat']
                            if row[27].value == None:
                                row[27].value = card['model']
                            if row[28].value == None:
                                row[28].value = card['brand']
                            if row[29].value != card['width']:
                                row[29].value = card['width']
                            cards.remove(card)
                            break
            else:
                del_rows.append(index+1)

        # remove old and empty rows 
        logging.info(f'<Record tires> The number of rows to be deleted: {len(del_rows)}')
        for i in reversed(del_rows):
            ws.delete_rows(i)

        if len(cards) > 0:
            logging.info(f"<Record tires> Last line number: {ws.max_row}")
            logging.info(f'<Record tires> The number of lines to be added {len(cards)}')
            for card in cards:
                index = ws.max_row + 1
                if ws[index][1].value == None:
                    ws[index][1].value = card['id']
                    ws[index][2].value = card['title']
                    ws[index][3].value = format_tire_desc(season=card['tire_type'], title=card['title'])
                    ws[index][4].value = card['price']
                    ws[index][5].value = card['diameter']
                    ws[index][6].value = 'Легковые шины'
                    photo_list = []
                    if 'live_photo' in card:
                        photo_list = ['https://rimzona.ru/media/{}'.format(slug) for slug in card['live_photo']] 
                    if 'photo' in card:
                        photo_list += ['https://rimzona.ru/media/{}'.format(slug) for slug in card['photo']]
                    ws[index][7].value = ' | '.join(photo_list)
                    ws[index][8].value = 'Шины, диски и колёса'
                    ws[index][9].value = 'Товар от производителя'
                    ws[index][10].value = random.choice(city_data['adress'])
                    ws[index][11].value = city_data['email']
                    ws[index][12].value = city_data['phone']
                    ws[index][13].value = '10-048'
                    ws[index][14].value = 'Новое'
                    ws[index][15].value = 'Активно'
                    ws[index][16].value = 'По телефону и в сообщениях'
                    ws[index][17].value = 'Запчасти и аксессуары'
                    ws[index][18].value = 'Package'
                    ws[index][19].value = 'Rimzona : сеть магазинов стильных дисков и шин'
                    ws[index][20].value = 'за 1 шт.'
                    ws[index][21].value = card['tire_type']
                    ws[index][22].value = card['height']
                    ws[index][23].value = card['load_index']
                    ws[index][24].value = 'Нет'
                    ws[index][25].value = card['speed_index']
                    ws[index][26].value = card['runflat']
                    ws[index][27].value = card['model']
                    ws[index][28].value = card['brand']
                    ws[index][29].value = card['width']
                    logging.debug(f"{index} : {ws[index][1].value} : {ws[index][9].value}")

        logging.info(f"<Record tires> -END- Last line number: {ws.max_row}")
        return 1
    else:
        raise Exception('Лист шины содержит некорректный титул')
